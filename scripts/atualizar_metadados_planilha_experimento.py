#!/usr/bin/env python3
"""Atualiza metadados confirmados após a conferência da coleta experimental."""

from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "planilha_coleta_rfid_uhf_experimento.xlsx"


def main() -> None:
    workbook = load_workbook(WORKBOOK)

    configuration = workbook["Configuracao"]
    configuration["B4"] = (
        "Laboratório de Práticas Digitais; ambiente fechado com área "
        "aproximada de 50 m²"
    )
    configuration["B15"] = (
        "Fonte externa regulada de 5 V/1 A e alimentação USB do Arduino"
    )
    configuration["C15"] = (
        "As duas formas de alimentação foram usadas; as tentativas não foram "
        "identificadas individualmente por fonte."
    )
    configuration["B20"] = (
        "Fotografias da montagem preservadas pelo autor; não incorporadas ao "
        "texto da monografia"
    )
    configuration["B21"] = (
        "SoftwareSerial a 115200 bit/s apresentou timeouts/quadros inválidos "
        "isolados. Execuções contaminadas por tags externas foram preservadas "
        "no log e repetidas. O inventário foi adaptado de 10 s para 25 s em "
        "todas as cinco rodadas. Após as medições com fonte externa de 5 V e "
        "alimentação USB do Arduino, não foi observada alteração relevante nos "
        "resultados em razão da forma de alimentação."
    )

    readme = workbook["LEIA_ME"]
    readme["A1"] = "REGISTRO DA CAMPANHA EXPERIMENTAL"
    readme["B3"] = (
        "Planilha preenchida com medições físicas reais realizadas em "
        "13/07/2026; resultados registrados durante a campanha."
    )
    readme["B4"] = (
        "Campanha reduzida concluída; o tempo de bancada limitou o número de "
        "repetições e de cenários."
    )
    readme["B5"] = (
        "Ordem executada: configuração/fotos, alcance, orientação, material e "
        "inventário. As condições foram agrupadas, sem aleatorização formal."
    )
    readme["B16"] = (
        "Foram executadas cinco janelas de 25 s com cinco tags e o mesmo "
        "percurso curto."
    )
    readme["B18"] = (
        "Arquivo conferido após a coleta; fotografias e registros seriais "
        "auxiliares foram preservados pelo autor."
    )

    plan = workbook["PLANO_DO_DIA"]
    plan["C7"] = (
        "Executar cinco janelas de inventário de 25 s com o mesmo percurso."
    )
    plan["E3"] = "X"
    plan["E8"] = "X"

    summary = workbook["Resumo"]
    summary["E2"] = "Detecções ou rodadas completas"
    summary["F2"] = "Taxa ou cobertura média (%)"

    tags = workbook["Tags"]
    for row in range(2, tags.max_row + 1):
        tags.cell(row, 6).value = (
            "Fotografia disponível e preservada pelo autor; posição registrada "
            "na planilha quando informada."
        )

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(WORKBOOK)


if __name__ == "__main__":
    main()
