#!/usr/bin/env python3
"""Consolida a planilha experimental do protótipo RFID UHF.

O script lê os registros individuais da campanha principal, incorpora as
contagens agregadas fornecidas para o ensaio complementar com uma etiqueta,
recalcula taxas e gera tabelas e figuras para a monografia. Os valores de RSSI
usados nas estatísticas são os bytes brutos anotados nas observações; eles não
são convertidos para dBm, pois a campanha não forneceu calibração para isso.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from math import sqrt
from pathlib import Path
from statistics import mean, stdev

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "planilha_coleta_rfid_uhf_experimento.xlsx"
TABLE_OUTPUT = ROOT / "05-tabelas" / "resultados_rfid_gerados.tex"
FIGURE_DIR = ROOT / "04-figuras"
TABLE_SOURCE = "Elaborado pelo autor a partir dos dados experimentais (2026)."
MATERIAL_LABELS = {
    "Papelao": "Papelão",
    "Metal direto": "Metal direto",
    "Metal com espacador": "Metal com espaçador",
}
MATERIAL_ALIASES = {
    "Papelao": "Papelao",
    "Papelão": "Papelao",
    "Metal direto": "Metal direto",
    "Metal com espacador": "Metal com espacador",
    "Metal com espaçador": "Metal com espacador",
}
TAG_SUPPORT_LABELS = {
    "TAG1": "TAG1 (papelão)",
    "TAG2": "TAG2 (madeira)",
    "TAG3": "TAG3 (plástico)",
}
COMPLEMENTARY_SINGLE_TAG_RESULTS = (
    # distância (m), orientação (graus), detecções, tentativas
    (1.0, 0, 30, 30),
    (1.0, 45, 26, 30),
    (1.0, 90, 18, 30),
    (2.0, 0, 27, 30),
    (2.0, 45, 21, 30),
    (2.0, 90, 11, 30),
    (5.0, 0, 0, 30),
    (5.0, 45, 0, 30),
    (5.0, 90, 0, 30),
)


def rows_as_dicts(workbook, sheet_name: str) -> list[dict[str, object]]:
    sheet = workbook[sheet_name]
    values = list(sheet.iter_rows(values_only=True))
    headers = [str(value) for value in values[0]]
    return [
        dict(zip(headers, row, strict=True))
        for row in values[1:]
        if any(value is not None for value in row)
    ]


def deduplicate_distance_rows(
    rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    """Remove repetições idênticas sem tratá-las como novas tentativas.

    A chave experimental é formada por distância, etiqueta e tentativa. O campo
    ``ordem`` é apenas um identificador de linha e, por isso, não integra a
    comparação dos resultados medidos. Se duas linhas com a mesma chave
    divergirem em qualquer outro campo, a consolidação é interrompida para
    impedir a escolha arbitrária de uma das observações.
    """
    unique_by_key: dict[
        tuple[float, str, int], dict[str, object]
    ] = {}
    deduplicated: list[dict[str, object]] = []
    discarded = 0

    for row in rows:
        key = (
            float(row["distancia_horizontal_m"]),
            str(row["tag_id"]),
            int(row["tentativa"]),
        )
        if key not in unique_by_key:
            unique_by_key[key] = row
            deduplicated.append(row)
            continue

        previous = unique_by_key[key]
        fields = (set(previous) | set(row)) - {"ordem"}
        divergent_fields = sorted(
            field for field in fields if previous.get(field) != row.get(field)
        )
        if divergent_fields:
            field_text = ", ".join(divergent_fields)
            raise ValueError(
                "linhas com a mesma chave experimental divergem: "
                f"distância={key[0]:g} m, etiqueta={key[1]}, "
                f"tentativa={key[2]}; campos: {field_text}"
            )
        discarded += 1

    if discarded:
        print(
            "aviso: "
            f"{discarded} linhas duplicadas exatas do ensaio de distância "
            "foram descartadas e não aumentaram o número de tentativas.",
            file=sys.stderr,
        )
    return deduplicated


def select_official_material_rows(
    rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    """Mantém somente as três condições validadas do ensaio de material."""
    selected: list[dict[str, object]] = []
    ignored: dict[str, int] = defaultdict(int)

    for row in rows:
        source_label = str(row["material_suporte"])
        canonical_label = MATERIAL_ALIASES.get(source_label)
        if canonical_label is None:
            ignored[source_label] += 1
            continue
        normalized_row = dict(row)
        normalized_row["material_suporte"] = canonical_label
        selected.append(normalized_row)

    if ignored:
        details = ", ".join(
            f"{label} ({count})" for label, count in sorted(ignored.items())
        )
        print(
            "aviso: condições adicionais de material sem validação "
            f"independente foram ignoradas: {details}.",
            file=sys.stderr,
        )

    present = {str(row["material_suporte"]) for row in selected}
    missing = set(MATERIAL_LABELS) - present
    if missing:
        missing_text = ", ".join(MATERIAL_LABELS[item] for item in sorted(missing))
        raise ValueError(
            "faltam condições oficiais no ensaio de material: " + missing_text
        )
    return selected


def pt_number(value: float, decimals: int = 1) -> str:
    return f"{value:.{decimals}f}".replace(".", ",")


def pt_optional_number(value: float | None, decimals: int = 1) -> str:
    if value is None:
        return "Sem leitura"
    return pt_number(value, decimals)


def mean_stdev_text(
    mean_value: float | None,
    stdev_value: float | None,
    decimals: int = 1,
) -> str:
    if mean_value is None:
        return "Sem leitura"
    if stdev_value is None:
        return pt_number(mean_value, decimals)
    return (
        f"{pt_number(mean_value, decimals)} $\\pm$ "
        f"{pt_number(stdev_value, decimals)}"
    )


def parse_metric(observation: object, key: str) -> int | None:
    if observation is None:
        return None
    match = re.search(rf"{re.escape(key)}=(\d+)", str(observation))
    if not match:
        return None
    return int(match.group(1))


def parse_rssi(observation: object) -> int | None:
    return parse_metric(observation, "rssi_raw")


def mean_or_none(values: list[float]) -> float | None:
    return mean(values) if values else None


def stdev_or_none(values: list[float]) -> float | None:
    return stdev(values) if len(values) > 1 else None


def wilson_interval(successes: int, attempts: int) -> tuple[float, float]:
    """Retorna o IC de Wilson de 95% para uma proporção binomial."""
    z = 1.959963984540054
    proportion = successes / attempts
    denominator = 1 + z**2 / attempts
    center = (proportion + z**2 / (2 * attempts)) / denominator
    half_width = (
        z
        * sqrt(
            proportion * (1 - proportion) / attempts
            + z**2 / (4 * attempts**2)
        )
        / denominator
    )
    return (center - half_width) * 100, (center + half_width) * 100


def mean_ci95(values: list[float]) -> tuple[float, float] | None:
    """IC t de 95% para a média; suficiente para as cinco rodadas atuais."""
    if len(values) < 2:
        return None
    critical_values = {
        2: 12.706,
        3: 4.303,
        4: 3.182,
        5: 2.776,
        6: 2.571,
        7: 2.447,
        8: 2.365,
        9: 2.306,
        10: 2.262,
    }
    critical = critical_values.get(len(values), 1.96)
    half_width = critical * stdev(values) / sqrt(len(values))
    return mean(values) - half_width, mean(values) + half_width


def aggregate_distance(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    groups: dict[float, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        groups[float(row["distancia_horizontal_m"])].append(row)

    result = []
    for distance, group in sorted(groups.items()):
        tag_groups: dict[str, list[dict[str, object]]] = defaultdict(list)
        for row in group:
            tag_groups[str(row["tag_id"])].append(row)

        attempts = len(group)
        valid = sum(int(row["detectou_0_1"]) for row in group)
        rates = [
            sum(int(row["detectou_0_1"]) for row in tag_rows) / len(tag_rows) * 100
            for tag_rows in tag_groups.values()
        ]
        first_reads = [
            float(row["tempo_primeira_leitura_s"])
            for row in group
            if row["tempo_primeira_leitura_s"] is not None
        ]
        rssi_values = [
            parse_rssi(row["observacoes"])
            for row in group
            if parse_rssi(row["observacoes"]) is not None
        ]

        result.append(
            {
                "distance": distance,
                "attempts": attempts,
                "valid": valid,
                "rate": valid / attempts * 100,
                "ci95": wilson_interval(valid, attempts),
                "minimum": min(rates),
                "maximum": max(rates),
                "first_time_mean": mean_or_none(first_reads),
                "first_time_stdev": stdev_or_none(first_reads),
                "rssi_mean": mean_or_none([float(value) for value in rssi_values]),
                "rssi_stdev": stdev_or_none([float(value) for value in rssi_values]),
            }
        )
    return result


def aggregate_orientation(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    groups: dict[float, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        groups[float(row["orientacao_graus"])].append(row)

    result = []
    for angle, group in sorted(groups.items()):
        tag_groups: dict[str, list[dict[str, object]]] = defaultdict(list)
        for row in group:
            tag_groups[str(row["tag_id"])].append(row)

        attempts = len(group)
        valid = sum(int(row["detectou_0_1"]) for row in group)
        rssi_values = [
            parse_rssi(row["observacoes"])
            for row in group
            if parse_rssi(row["observacoes"]) is not None
        ]
        result.append(
            {
                "angle": angle,
                "attempts": attempts,
                "valid": valid,
                "rate": valid / attempts * 100,
                "ci95": wilson_interval(valid, attempts),
                "minimum": min(
                    sum(int(row["detectou_0_1"]) for row in tag_rows)
                    / len(tag_rows)
                    * 100
                    for tag_rows in tag_groups.values()
                ),
                "maximum": max(
                    sum(int(row["detectou_0_1"]) for row in tag_rows)
                    / len(tag_rows)
                    * 100
                    for tag_rows in tag_groups.values()
                ),
                "rssi_mean": mean_or_none([float(value) for value in rssi_values]),
                "rssi_stdev": stdev_or_none([float(value) for value in rssi_values]),
            }
        )
    return result


def aggregate_material(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    groups: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        groups[str(row["material_suporte"])].append(row)

    result = []
    for material, group in groups.items():
        attempts = len(group)
        valid = sum(int(row["detectou_0_1"]) for row in group)
        rssi_values = [
            parse_rssi(row["observacoes"])
            for row in group
            if parse_rssi(row["observacoes"]) is not None
        ]
        result.append(
            {
                "material": material,
                "attempts": attempts,
                "valid": valid,
                "rate": valid / attempts * 100,
                "ci95": wilson_interval(valid, attempts),
                "rssi_mean": mean_or_none([float(value) for value in rssi_values]),
                "rssi_stdev": stdev_or_none([float(value) for value in rssi_values]),
            }
        )

    order = {
        "Papelao": 0,
        "Metal direto": 1,
        "Metal com espacador": 2,
    }
    result.sort(key=lambda row: order.get(str(row["material"]), 99))
    return result


def inventory_coverage(row: dict[str, object]) -> float:
    """Calcula a cobertura sem depender do cache de fórmulas do Excel."""
    present = int(row["qtd_tags_presentes"])
    detected = int(row["qtd_tags_detectadas"])
    if present <= 0:
        raise ValueError("qtd_tags_presentes deve ser maior que zero")
    return detected / present * 100


def inventory_complete(row: dict[str, object]) -> int:
    """Indica inventário completo a partir das contagens registradas."""
    return int(
        int(row["qtd_tags_detectadas"]) >= int(row["qtd_tags_presentes"])
    )


def inventory_metrics(rows: list[dict[str, object]]) -> dict[str, object]:
    coverages = [inventory_coverage(row) for row in rows]
    duplicates = [
        parse_metric(row["observacoes"], "duplicatas")
        for row in rows
        if parse_metric(row["observacoes"], "duplicatas") is not None
    ]
    valid_reads = [
        parse_metric(row["observacoes"], "leituras_validas")
        for row in rows
        if parse_metric(row["observacoes"], "leituras_validas") is not None
    ]
    coverage_ci95 = mean_ci95(coverages)
    complete_rounds = sum(inventory_complete(row) for row in rows)
    complete_rate = complete_rounds / len(rows) * 100
    return {
        "rounds": len(rows),
        "tags": int(rows[0]["qtd_tags_presentes"]),
        "coverages": coverages,
        "mean_coverage": mean(coverages),
        "stdev_coverage": stdev(coverages),
        "coverage_ci95": coverage_ci95,
        "minimum_coverage": min(coverages),
        "maximum_coverage": max(coverages),
        "complete_rounds": complete_rounds,
        "complete_rate": complete_rate,
        "complete_ci95": wilson_interval(complete_rounds, len(rows)),
        "external_reads": sum(int(row["leituras_externas"]) for row in rows),
        "mean_duplicates": mean(duplicates) if duplicates else 0.0,
        "total_duplicates": sum(duplicates) if duplicates else 0,
        "mean_valid_reads": mean(valid_reads) if valid_reads else 0.0,
    }


def tag_performance(rows: list[dict[str, object]]) -> dict[str, dict[str, float | int]]:
    groups: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        groups[str(row["tag_id"])].append(row)
    return {
        tag: {
            "attempts": len(group),
            "valid": sum(int(row["detectou_0_1"]) for row in group),
            "rate": sum(int(row["detectou_0_1"]) for row in group)
            / len(group)
            * 100,
        }
        for tag, group in sorted(groups.items())
    }


def render_table(
    distance: list[dict[str, object]],
    orientation: list[dict[str, object]],
    material: list[dict[str, object]],
    inventory: dict[str, object],
    distance_tags: dict[str, dict[str, float | int]],
    orientation_tags: dict[str, dict[str, float | int]],
) -> str:
    distance_rows = "\n".join(
        "        "
        f"{pt_number(float(row['distance']))} & {row['valid']}/{row['attempts']} & "
        f"{pt_number(float(row['rate']))}\\% & "
        f"{pt_number(float(row['ci95'][0]))}\\% a "
        f"{pt_number(float(row['ci95'][1]))}\\% & "
        f"{mean_stdev_text(row['first_time_mean'], row['first_time_stdev'], 3)} & "
        f"{mean_stdev_text(row['rssi_mean'], row['rssi_stdev'])} \\\\"
        for row in distance
    )
    orientation_rows = "\n".join(
        "        "
        f"{int(row['angle'])} & {row['valid']}/{row['attempts']} & "
        f"{pt_number(float(row['rate']))}\\% & "
        f"{pt_number(float(row['ci95'][0]))}\\% a "
        f"{pt_number(float(row['ci95'][1]))}\\% & "
        f"{mean_stdev_text(row['rssi_mean'], row['rssi_stdev'])} \\\\"
        for row in orientation
    )
    material_rows = "\n".join(
        "        "
        f"{MATERIAL_LABELS.get(str(row['material']), str(row['material']))} & "
        f"{row['valid']}/{row['attempts']} & {pt_number(float(row['rate']))}\\% & "
        f"{pt_number(float(row['ci95'][0]))}\\% a "
        f"{pt_number(float(row['ci95'][1]))}\\% & "
        f"{mean_stdev_text(row['rssi_mean'], row['rssi_stdev'])} \\\\"
        for row in material
    )
    tag_rows = "\n".join(
        "        "
        f"{TAG_SUPPORT_LABELS.get(tag, tag)} & "
        f"{distance_tags[tag]['valid']}/{distance_tags[tag]['attempts']} & "
        f"{pt_number(float(distance_tags[tag]['rate']))}\\% & "
        f"{orientation_tags[tag]['valid']}/{orientation_tags[tag]['attempts']} & "
        f"{pt_number(float(orientation_tags[tag]['rate']))}\\% \\\\"
        for tag in distance_tags
    )
    complementary_rows = "\n".join(
        "        "
        f"{pt_number(distance)} & {angle} & {successes}/{attempts} & "
        f"{pt_number(successes / attempts * 100)}\\% & "
        f"{pt_number(wilson_interval(successes, attempts)[0])}\\% a "
        f"{pt_number(wilson_interval(successes, attempts)[1])}\\% \\\\"
        for distance, angle, successes, attempts
        in COMPLEMENTARY_SINGLE_TAG_RESULTS
    )
    inventory_rows = "\n".join(
        "        "
        f"{index} & {int(row['qtd_tags_detectadas'])} & {pt_number(inventory_coverage(row))}\\% & "
        f"{'Sim' if inventory_complete(row) else 'Não'} & {int(row['leituras_externas'])} \\\\"
        for index, row in enumerate(inventory["rows"], start=1)
    )

    return f"""% Tabelas elaboradas a partir dos dados experimentais coletados.
% Os níveis de RSSI são apresentados na unidade bruta informada pelo módulo.

\\newcommand{{\\TabelaAlcanceDistancia}}{{%
\\begin{{table}}[htbp]
    \\centering
    \\caption{{Resultado experimental por distância horizontal.}}
    \\label{{tab:resultado-alcance-distancia}}
    \\footnotesize
    \\setlength{{\\tabcolsep}}{{3.5pt}}
    \\begin{{tabular}}{{cccccc}}
        \\toprule
        \\textbf{{\\shortstack{{Distância\\\\(m)}}}} &
        \\textbf{{\\shortstack{{Detecções\\\\($n/N$)}}}} &
        \\textbf{{\\shortstack{{Taxa de\\\\leitura}}}} &
        \\textbf{{IC 95\\%}} &
        \\textbf{{\\shortstack{{Tempo da primeira leitura\\\\média $\\pm$ DP (s)}}}} &
        \\textbf{{\\shortstack{{RSSI bruto\\\\média $\\pm$ DP}}}} \\\\
        \\midrule
{distance_rows}
        \\bottomrule
    \\end{{tabular}}
    \\fonte{{{TABLE_SOURCE}}}
\\end{{table}}%
}}

\\newcommand{{\\TabelaOrientacao}}{{%
\\begin{{table}}[htbp]
    \\centering
    \\caption{{Resultado experimental por orientação da etiqueta.}}
    \\label{{tab:resultado-orientacao}}
    \\small
    \\setlength{{\\tabcolsep}}{{5pt}}
    \\begin{{tabular}}{{ccccc}}
        \\toprule
        \\textbf{{\\shortstack{{Orientação\\\\($^\\circ$)}}}} &
        \\textbf{{\\shortstack{{Detecções\\\\($n/N$)}}}} &
        \\textbf{{\\shortstack{{Taxa de\\\\leitura}}}} &
        \\textbf{{IC 95\\%}} &
        \\textbf{{\\shortstack{{RSSI bruto\\\\média $\\pm$ DP}}}} \\\\
        \\midrule
{orientation_rows}
        \\bottomrule
    \\end{{tabular}}
    \\fonte{{{TABLE_SOURCE}}}
\\end{{table}}%
}}

\\newcommand{{\\TabelaMaterial}}{{%
\\begin{{table}}[htbp]
    \\centering
    \\caption{{Resultado experimental por material de fixação.}}
    \\label{{tab:resultado-material}}
    \\small
    \\setlength{{\\tabcolsep}}{{4pt}}
    \\begin{{tabular}}{{lcccc}}
        \\toprule
        \\textbf{{Material de fixação}} &
        \\textbf{{\\shortstack{{Detecções\\\\($n/N$)}}}} &
        \\textbf{{\\shortstack{{Taxa de\\\\leitura}}}} &
        \\textbf{{IC 95\\%}} &
        \\textbf{{\\shortstack{{RSSI bruto\\\\média $\\pm$ DP}}}} \\\\
        \\midrule
{material_rows}
        \\bottomrule
    \\end{{tabular}}
    \\fonte{{{TABLE_SOURCE}}}
\\end{{table}}%
}}

\\newcommand{{\\TabelaPorEtiqueta}}{{%
\\begin{{table}}[htbp]
    \\centering
    \\caption{{Desempenho agregado por conjunto formado pela etiqueta e pelo suporte nos ensaios de distância e orientação.}}
    \\label{{tab:resultado-por-tag}}
    \\small
    \\setlength{{\\tabcolsep}}{{4pt}}
    \\begin{{tabular}}{{lrrrr}}
        \\toprule
        \\textbf{{\\shortstack{{Etiqueta\\\\e suporte}}}} &
        \\textbf{{\\shortstack{{Distância\\\\detecções ($n/N$)}}}} &
        \\textbf{{\\shortstack{{Taxa de\\\\leitura}}}} &
        \\textbf{{\\shortstack{{Orientação\\\\detecções ($n/N$)}}}} &
        \\textbf{{\\shortstack{{Taxa de\\\\leitura}}}} \\\\
        \\midrule
{tag_rows}
        \\bottomrule
    \\end{{tabular}}
    \\fonte{{{TABLE_SOURCE}}}
\\end{{table}}%
}}

\\newcommand{{\\TabelaEnsaioUmaTag}}{{%
\\begin{{table}}[htbp]
    \\centering
    \\caption{{Ensaio complementar com uma única etiqueta e 30 tentativas por combinação.}}
    \\label{{tab:resultado-ensaio-uma-tag}}
    \\small
    \\setlength{{\\tabcolsep}}{{5pt}}
    \\begin{{tabular}}{{ccccc}}
        \\toprule
        \\textbf{{\\shortstack{{Distância\\\\(m)}}}} &
        \\textbf{{\\shortstack{{Orientação\\\\($^\\circ$)}}}} &
        \\textbf{{\\shortstack{{Detecções\\\\($n/N$)}}}} &
        \\textbf{{\\shortstack{{Taxa de\\\\leitura}}}} &
        \\textbf{{IC 95\\% de Wilson}} \\\\
        \\midrule
{complementary_rows}
        \\bottomrule
    \\end{{tabular}}
    \\fonte{{{TABLE_SOURCE}}}
\\end{{table}}%
}}

\\newcommand{{\\TabelaInventarioRodadas}}{{%
\\begin{{table}}[htbp]
    \\centering
    \\caption{{Cobertura por rodada no inventário experimental em sala.}}
    \\label{{tab:resultado-inventario-rodadas}}
    \\small
    \\setlength{{\\tabcolsep}}{{6pt}}
    \\begin{{tabular}}{{ccccc}}
        \\toprule
        \\textbf{{Rodada}} &
        \\textbf{{\\shortstack{{Etiquetas\\\\detectadas}}}} &
        \\textbf{{Cobertura}} &
        \\textbf{{\\shortstack{{Rodada\\\\completa}}}} &
        \\textbf{{\\shortstack{{Leituras\\\\externas}}}} \\\\
        \\midrule
{inventory_rows}
        \\bottomrule
    \\end{{tabular}}
    \\fonte{{{TABLE_SOURCE}}}
\\end{{table}}%
}}

\\newcommand{{\\TabelaInventarioResumo}}{{%
\\begin{{table}}[htbp]
    \\centering
    \\caption{{Resumo do inventário experimental.}}
    \\label{{tab:resultado-inventario}}
    \\small
    \\begin{{tabular}}{{p{{6.2cm}}r}}
        \\toprule
        \\textbf{{Métrica}} & \\textbf{{Valor}} \\\\
        \\midrule
        Rodadas analisadas & {inventory['rounds']} \\\\
        Etiquetas presentes por rodada & {inventory['tags']} \\\\
        Cobertura média & {pt_number(float(inventory['mean_coverage']))}\\% \\\\
        Desvio-padrão da cobertura & {pt_number(float(inventory['stdev_coverage']))} pontos percentuais \\\\
        IC 95\\% da cobertura média & {pt_number(float(inventory['coverage_ci95'][0]))}\\% a {pt_number(float(inventory['coverage_ci95'][1]))}\\% \\\\
        Faixa de cobertura & {pt_number(float(inventory['minimum_coverage']))}\\% a {pt_number(float(inventory['maximum_coverage']))}\\% \\\\
        Taxa de rodadas completas & {inventory['complete_rounds']}/{inventory['rounds']} = {pt_number(float(inventory['complete_rate']))}\\% \\\\
        IC 95\\% da taxa de rodadas completas & {pt_number(float(inventory['complete_ci95'][0]))}\\% a {pt_number(float(inventory['complete_ci95'][1]))}\\% \\\\
        Leituras externas totais & {inventory['external_reads']} \\\\
        Leituras repetidas totais & {inventory['total_duplicates']} \\\\
        Leituras repetidas médias por rodada & {pt_number(float(inventory['mean_duplicates']))} \\\\
        Leituras válidas médias por rodada & {pt_number(float(inventory['mean_valid_reads']))} \\\\
        \\bottomrule
    \\end{{tabular}}
    \\fonte{{{TABLE_SOURCE}}}
\\end{{table}}%
}}
"""


def save_figure(figure, filename: str) -> None:
    path = FIGURE_DIR / filename
    figure.savefig(path, dpi=170, bbox_inches="tight", facecolor="white")


def write_figures(
    distance: list[dict[str, object]],
    orientation: list[dict[str, object]],
    material: list[dict[str, object]],
    inventory: dict[str, object],
) -> None:
    import matplotlib.pyplot as plt

    plt.rcParams.update(
        {
            "font.size": 12,
            "axes.labelsize": 14,
            "xtick.labelsize": 12,
            "ytick.labelsize": 12,
            "axes.grid": True,
            "grid.alpha": 0.25,
        }
    )

    x = [float(row["distance"]) for row in distance]
    y = [float(row["rate"]) for row in distance]
    lower = [value - float(row["ci95"][0]) for value, row in zip(y, distance)]
    upper = [float(row["ci95"][1]) - value for value, row in zip(y, distance)]
    fig, ax = plt.subplots(figsize=(10, 5.6), constrained_layout=True)
    ax.errorbar(
        x,
        y,
        yerr=[lower, upper],
        marker="o",
        linewidth=2.6,
        capsize=5,
        color="#24658f",
    )
    ax.set(
        xlabel="Distância horizontal entre antena e etiqueta (m)",
        ylabel="Taxa de leitura (%)",
        ylim=(0, 105),
    )
    ax.set_xticks(x, [pt_number(value) for value in x])
    save_figure(fig, "resultado_alcance_distancia.png")
    plt.close(fig)

    time_rows = [
        row
        for row in distance
        if row["first_time_mean"] is not None
    ]
    tx = [float(row["distance"]) for row in time_rows]
    ty = [float(row["first_time_mean"]) for row in time_rows]
    terr = [float(row["first_time_stdev"] or 0.0) for row in time_rows]
    fig, ax = plt.subplots(figsize=(10, 5.6), constrained_layout=True)
    ax.errorbar(
        tx,
        ty,
        yerr=terr,
        marker="o",
        linewidth=2.6,
        capsize=5,
        color="#813d18",
    )
    ax.set(
        xlabel="Distância horizontal entre antena e etiqueta (m)",
        ylabel="Tempo médio até a primeira leitura (s)",
        ylim=(0, None),
    )
    save_figure(fig, "resultado_tempo_primeira_leitura.png")
    plt.close(fig)

    angles = [int(row["angle"]) for row in orientation]
    angle_rates = [float(row["rate"]) for row in orientation]
    angle_lower = [
        value - float(row["ci95"][0])
        for value, row in zip(angle_rates, orientation)
    ]
    angle_upper = [
        float(row["ci95"][1]) - value
        for value, row in zip(angle_rates, orientation)
    ]
    fig, ax = plt.subplots(figsize=(8.5, 5.6), constrained_layout=True)
    ax.bar(
        [str(value) for value in angles],
        angle_rates,
        yerr=[angle_lower, angle_upper],
        capsize=5,
        color="#348663",
    )
    ax.set(
        xlabel="Orientação da etiqueta (graus)",
        ylabel="Taxa de leitura (%)",
        ylim=(0, 105),
    )
    save_figure(fig, "resultado_orientacao.png")
    plt.close(fig)

    materials = [
        MATERIAL_LABELS.get(str(row["material"]), str(row["material"]))
        for row in material
    ]
    material_rates = [float(row["rate"]) for row in material]
    material_lower = [
        value - float(row["ci95"][0])
        for value, row in zip(material_rates, material)
    ]
    material_upper = [
        float(row["ci95"][1]) - value
        for value, row in zip(material_rates, material)
    ]
    fig, ax = plt.subplots(figsize=(10, 5.8), constrained_layout=True)
    ax.bar(
        materials,
        material_rates,
        yerr=[material_lower, material_upper],
        capsize=5,
        color="#7965a8",
    )
    ax.set(
        xlabel="Material de fixação",
        ylabel="Taxa de leitura (%)",
        ylim=(0, 105),
    )
    ax.tick_params(axis="x", rotation=18)
    save_figure(fig, "resultado_material.png")
    plt.close(fig)

    rounds = list(range(1, len(inventory["rows"]) + 1))
    fig, ax = plt.subplots(figsize=(10, 5.6), constrained_layout=True)
    ax.bar(rounds, inventory["coverages"], color="#b86139")
    ax.set(
        xlabel="Rodada de inventário",
        ylabel="Cobertura (%)",
        ylim=(0, 105),
        xticks=rounds,
    )
    save_figure(fig, "resultado_inventario.png")
    plt.close(fig)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--write",
        action="store_true",
        help="grava a tabela LaTeX e as figuras",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    if not WORKBOOK.is_file():
        print(f"erro: arquivo não encontrado: {WORKBOOK}", file=sys.stderr)
        return 2

    # As fórmulas existentes na planilha são deliberadamente ignoradas. Todos
    # os indicadores são recalculados abaixo a partir das contagens brutas.
    workbook = load_workbook(WORKBOOK, data_only=False)
    distance_rows = rows_as_dicts(workbook, "Alcance")
    orientation_rows = rows_as_dicts(workbook, "Orientacao")
    material_rows = rows_as_dicts(workbook, "Material")
    inventory_rows = rows_as_dicts(workbook, "Inventario")

    distance_rows = deduplicate_distance_rows(distance_rows)
    material_rows = select_official_material_rows(material_rows)
    distance = aggregate_distance(distance_rows)
    orientation = aggregate_orientation(orientation_rows)
    material = aggregate_material(material_rows)
    inventory = inventory_metrics(inventory_rows)
    inventory["rows"] = inventory_rows
    distance_tags = tag_performance(distance_rows)
    orientation_tags = tag_performance(orientation_rows)

    print("Métricas experimentais recalculadas:")
    for row in distance:
        print(
            f"  distância {row['distance']:.1f} m: "
            f"{row['valid']}/{row['attempts']} = {row['rate']:.1f}%"
        )
    for row in orientation:
        print(
            f"  orientação {row['angle']:.0f}°: "
            f"{row['valid']}/{row['attempts']} = {row['rate']:.1f}%"
        )
    print(
        f"  inventário: cobertura média {inventory['mean_coverage']:.1f}%; "
        f"{inventory['complete_rounds']}/{inventory['rounds']} rodadas completas"
    )

    if not args.write:
        print("Verificação concluída sem alterar arquivos.")
        return 0

    TABLE_OUTPUT.write_text(
        render_table(
            distance,
            orientation,
            material,
            inventory,
            distance_tags,
            orientation_tags,
        ),
        encoding="utf-8",
        newline="\n",
    )
    write_figures(distance, orientation, material, inventory)
    print(f"Tabelas gravadas em: {TABLE_OUTPUT.resolve()}")
    print(f"Figuras gravadas em: {FIGURE_DIR.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
