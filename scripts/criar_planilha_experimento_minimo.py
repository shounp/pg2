#!/usr/bin/env python3
"""Gera uma planilha limpa para uma campanha RFID UHF mínima e real."""

from __future__ import annotations

import argparse
from pathlib import Path
import random
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "planilha_coleta_rfid_uhf_experimento.xlsx"

NAVY = "1F4E78"
BLUE = "5B9BD5"
LIGHT_BLUE = "DDEBF7"
YELLOW = "FFF2CC"
GREEN = "E2F0D9"
RED = "F4CCCC"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="B7B7B7")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="substitui somente a planilha experimental gerada anteriormente",
    )
    return parser.parse_args()


def style_header(sheet, row: int = 1) -> None:
    for cell in sheet[row]:
        if cell.value is None:
            continue
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
    sheet.row_dimensions[row].height = 34


def style_header_range(sheet, row: int, first_column: int, last_column: int) -> None:
    for column in range(first_column, last_column + 1):
        cell = sheet.cell(row, column)
        if cell.value is None:
            continue
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)


def style_title(sheet, title: str, last_column: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    cell = sheet.cell(1, 1, title)
    cell.font = Font(size=16, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28


def set_widths(sheet, widths: list[float]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def mark_input(cell, comment: str | None = None) -> None:
    cell.fill = PatternFill("solid", fgColor=YELLOW)
    if comment:
        cell.comment = Comment(comment, "Protocolo PG2")


def mark_formula(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)


def add_binary_validation(sheet, cell_range: str) -> None:
    validation = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    validation.error = "Digite somente 0 ou 1."
    validation.errorTitle = "Valor inválido"
    validation.prompt = "1 = detectou; 0 = não detectou."
    validation.promptTitle = "Resultado da tentativa"
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def add_list_validation(sheet, cell_range: str, values: list[object]) -> None:
    """Adiciona uma lista curta de valores permitidos, aceitando célula vazia."""
    formula = '"' + ",".join(str(value) for value in values) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "Selecione um dos valores previstos no protocolo."
    validation.errorTitle = "Valor inválido"
    validation.prompt = "Use a lista para não criar níveis experimentais acidentais."
    validation.promptTitle = "Condição planejada"
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def add_decimal_validation(
    sheet,
    cell_range: str,
    minimum: float = 0,
    maximum: float = 1_000_000,
) -> None:
    validation = DataValidation(
        type="decimal",
        operator="between",
        formula1=str(minimum),
        formula2=str(maximum),
        allow_blank=True,
    )
    validation.error = f"Digite um número entre {minimum} e {maximum}."
    validation.errorTitle = "Valor inválido"
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def add_whole_validation(
    sheet,
    cell_range: str,
    minimum: int = 0,
    maximum: int = 1_000_000,
) -> None:
    validation = DataValidation(
        type="whole",
        operator="between",
        formula1=str(minimum),
        formula2=str(maximum),
        allow_blank=True,
    )
    validation.error = f"Digite um número inteiro entre {minimum} e {maximum}."
    validation.errorTitle = "Valor inválido"
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def add_header_comments(sheet, comments: dict[str, str]) -> None:
    for coordinate, text in comments.items():
        sheet[coordinate].comment = Comment(text, "Protocolo PG2")


def style_side_title(sheet, cell_range: str, title: str) -> None:
    sheet.merge_cells(cell_range)
    first = cell_range.split(":", maxsplit=1)[0]
    cell = sheet[first]
    cell.value = title
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


FUTURE_CAMPAIGN_README_ROWS = [
    (
        "Fatorial futuro",
        "A aba Fatorial_Tag_Suporte é um plano de campanha futura/não realizada. "
        "As 162 linhas não contêm resultados; remonte cada combinação tag–suporte em cada bloco.",
    ),
    (
        "Repetição futura — S2",
        "A aba Repeticao_Sessao2 é um plano de campanha futura/não realizada. Primeiro repita "
        "em outro dia no mesmo laboratório. Um segundo ambiente é opcional; não troque dia e "
        "ambiente simultaneamente se a intenção for separar estabilidade temporal do efeito da sala.",
    ),
    (
        "Antenas — campanha futura",
        "A aba Comparacao_Antenas é um plano de campanha futura/não realizada. Prefira EIRP "
        "equivalente. Em antena dual-port, registre a porta/modo: uma porta isolada não constitui diversidade.",
    ),
]


def create_readme(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "LEIA_ME"
    style_title(sheet, "PROTOCOLO EXPERIMENTAL MÍNIMO — PREENCHIMENTO OBRIGATÓRIO", 2)
    rows = [
        ("Natureza", "Planilha vazia para coleta física real. Nenhum resultado foi pré-preenchido."),
        ("Tempo total", "Planejada para aproximadamente 60 a 80 minutos, incluindo montagem, fotos e conferência."),
        ("Prioridade", "Execute nesta ordem: Configuração/fotos, Alcance, Orientação, Material e Inventário."),
        ("Tentativa", "Uma janela de 3 s. Retire a tag do campo por cerca de 2 s antes de cada nova tentativa."),
        ("Sucesso", "Preencha 1 somente se o EPC esperado aparecer ao menos uma vez durante a janela; caso contrário, 0."),
        ("Não inventar", "Não complete células esquecidas posteriormente por estimativa. Deixe em branco e explique em observações."),
        ("RSSI", "Não é obrigatório. Só registre se o firmware fornecer RSSI bruto ou uma conversão documentada."),
        ("Cor amarela", "Campo que deve ser preenchido durante ou imediatamente após a coleta."),
        ("Cor azul", "Célula calculada automaticamente; não digitar por cima."),
        ("Tags", "Cadastre cinco tags. TAG1, TAG2 e TAG3 são usadas nos testes controlados; as cinco entram no inventário."),
        ("Alcance", "4 distâncias × 3 tags × 5 tentativas = 60 janelas de 3 s."),
        ("Orientação", "3 ângulos × 3 tags × 5 tentativas = 45 janelas de 3 s, sempre a 1,5 m."),
        ("Material", "6 condições × 5 tentativas = 30 janelas de 3 s, usando sempre TAG1."),
        ("Inventário", "5 janelas de 25 s com cinco tags e o mesmo percurso curto."),
        ("Passagem", "Foi retirada do protocolo para reduzir tempo e preservar a qualidade dos quatro ensaios centrais."),
        ("Backup", "Ao terminar, salve uma cópia com data/hora e preserve fotografias e qualquer log serial."),
    ]
    sheet.append(["Campo", "Instrução"])
    style_header(sheet, 2)
    for label, instruction in rows:
        sheet.append([label, instruction])
    for row in range(3, sheet.max_row + 1):
        sheet.cell(row, 1).font = Font(bold=True)
        sheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row].height = 32
    set_widths(sheet, [24, 105])
    sheet.freeze_panes = "A3"
    sheet.sheet_view.showGridLines = False
    add_future_campaign_readme_entries(workbook)


def add_future_campaign_readme_entries(workbook: Workbook) -> None:
    """Acrescenta, sem duplicar, o aviso de que as novas abas ainda não têm coleta."""
    sheet = workbook["LEIA_ME"]
    existing_labels = {
        sheet.cell(row, 1).value
        for row in range(1, sheet.max_row + 1)
        if sheet.cell(row, 1).value is not None
    }
    for label, instruction in FUTURE_CAMPAIGN_README_ROWS:
        if label in existing_labels:
            continue
        sheet.append([label, instruction])
        row = sheet.max_row
        sheet.cell(row, 1).font = Font(bold=True, color=NAVY)
        sheet.cell(row, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        sheet.cell(row, 2).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        sheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row].height = 46


def create_day_plan(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("PLANO_DO_DIA")
    style_title(sheet, "ROTEIRO RÁPIDO DE EXECUÇÃO", 5)
    headers = ["Etapa", "Tempo", "O que fazer", "Critério para avançar", "Concluído (X)"]
    sheet.append(headers)
    style_header(sheet, 2)
    rows = [
        (1, "10 min", "Montar, conferir GND/fonte, abrir serial, cadastrar 5 EPCs e tirar fotos.", "Leitura de TAG1 confirmada e configuração preenchida."),
        (2, "20 min", "Executar as 60 janelas da aba Alcance, agrupadas por distância.", "Todas as células amarelas de detectou_0_1 preenchidas."),
        (3, "12 min", "Executar as 45 janelas da aba Orientacao em 0°, 45° e 90°.", "Ângulos fotografados ou marcados e resultados preenchidos."),
        (4, "15 min", "Executar papelão, plástico, madeira, vidro, metal direto e metal com espaçador usando TAG1.", "Cinco tentativas preenchidas em cada uma das seis condições."),
        (5, "10 min", "Executar cinco janelas de inventário de 25 s com o mesmo percurso.", "EPCs detectados e quantidade registrados em cada janela."),
        (6, "5 min", "Revisar campos vazios, salvar cópia e copiar fotos/logs.", "Arquivo abre novamente e a aba Resumo apresenta taxas."),
    ]
    for item in rows:
        sheet.append(item)
        mark_input(sheet.cell(sheet.max_row, 5))
    for row in range(3, sheet.max_row + 1):
        for column in range(1, 6):
            sheet.cell(row, column).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row].height = 48
    set_widths(sheet, [10, 12, 66, 55, 16])
    sheet.freeze_panes = "A3"
    sheet.sheet_view.showGridLines = False


def create_configuration(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Configuracao")
    sheet.append(["Parâmetro", "Valor real", "Como preencher"])
    style_header(sheet)
    rows = [
        ("Data da coleta", None, "Data em que os experimentos foram realmente executados."),
        ("Hora de início", None, "Hora aproximada do início da montagem."),
        ("Local/sala", None, "Nome exato da sala e prédio."),
        ("Operador", "Leonardo Teixeira Gomes Silva", "Pessoa que realizou a coleta."),
        ("Leitor", "YPD-R200", "Confirmar a identificação do módulo."),
        ("Placa", None, "Arduino Uno ou Nano; informar qual foi usado."),
        ("Firmware/commit", None, "Nome/versão do arquivo carregado no Arduino."),
        ("Antena", "Airplux APCA8090", "Confirmar modelo/variante disponível."),
        ("Ganho nominal (dBi)", 5, "Valor nominal do fabricante."),
        ("Polarização", "Linear", "Polarização informada para a antena."),
        ("Região/canal/FHSS", None, "Registrar o que estiver configurado; se desconhecido, escrever NÃO VERIFICADO."),
        ("Potência configurada (dBm)", 26, "Registrar o comando usado; indicar se foi lido de volta."),
        ("Altura da antena (cm)", 150, "Medir do piso ao centro da antena."),
        ("Fonte do leitor", "5 V / 1 A externa", "Registrar tensão/corrente nominais da fonte."),
        ("Baud rate UART", 115200, "Velocidade entre Arduino e YPD-R200."),
        ("Intervalo de estabilização", None, "Tempo aguardado após mover tag/antena, em segundos."),
        ("Pessoas no ambiente", None, "Quantidade aproximada e movimentação durante o teste."),
        ("Objetos metálicos próximos", None, "Descrever bancada, armários e distância aproximada."),
        ("Fotos/arquivos", None, "Nomes das fotos e do log serial, se houver."),
        ("Ocorrências", None, "Reinicializações, quedas de alimentação ou mudanças de configuração."),
    ]
    for parameter, value, instruction in rows:
        sheet.append([parameter, value, instruction])
        mark_input(sheet.cell(sheet.max_row, 2))
        sheet.cell(sheet.max_row, 3).alignment = Alignment(wrap_text=True, vertical="top")
    set_widths(sheet, [32, 38, 72])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:C{sheet.max_row}"


def create_tags(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Tags")
    headers = ["tag_id", "epc_real", "objeto/suporte", "material", "foto/posição", "observações"]
    sheet.append(headers)
    style_header(sheet)
    for index in range(1, 6):
        sheet.append([f"TAG{index}", None, None, None, None, None])
        for column in range(2, 7):
            mark_input(sheet.cell(sheet.max_row, column))
    set_widths(sheet, [12, 42, 28, 22, 28, 55])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:F6"


def create_repeated_test_sheet(
    workbook: Workbook,
    name: str,
    headers: list[str],
    rows: list[list[object]],
    input_columns: list[int],
    widths: list[float],
    binary_column: int,
) -> None:
    sheet = workbook.create_sheet(name)
    sheet.append(headers)
    style_header(sheet)
    for values in rows:
        sheet.append(values)
        for column in input_columns:
            mark_input(sheet.cell(sheet.max_row, column))
    last_column = get_column_letter(len(headers))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"
    set_widths(sheet, widths)
    add_binary_validation(sheet, f"{get_column_letter(binary_column)}2:{get_column_letter(binary_column)}{sheet.max_row}")
    result_range = f"{get_column_letter(binary_column)}2:{get_column_letter(binary_column)}{sheet.max_row}"
    sheet.conditional_formatting.add(
        result_range,
        CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor=RED)),
    )
    sheet.conditional_formatting.add(
        result_range,
        CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor=GREEN)),
    )


def create_distance(workbook: Workbook) -> None:
    headers = [
        "ordem", "distancia_horizontal_m", "tag_id", "tentativa", "altura_antena_cm",
        "altura_tag_cm", "orientacao_graus", "material_suporte", "duracao_janela_s",
        "detectou_0_1", "epc_observado", "tempo_primeira_leitura_s", "observacoes",
    ]
    rows = []
    order = 1
    for distance in (0.5, 1.0, 1.5, 2.0):
        for tag in ("TAG1", "TAG2", "TAG3"):
            for attempt in range(1, 6):
                rows.append([order, distance, tag, attempt, 150, 150, 0, "Papelao", 3, None, None, None, None])
                order += 1
    create_repeated_test_sheet(
        workbook, "Alcance", headers, rows, [10, 11, 12, 13],
        [9, 20, 11, 10, 18, 16, 18, 21, 18, 16, 38, 24, 55], 10,
    )


def create_orientation(workbook: Workbook) -> None:
    headers = [
        "ordem", "orientacao_graus", "tag_id", "tentativa", "distancia_m",
        "altura_antena_cm", "altura_tag_cm", "material_suporte", "duracao_janela_s",
        "detectou_0_1", "epc_observado", "tempo_primeira_leitura_s", "observacoes",
    ]
    rows = []
    order = 1
    for angle in (0, 45, 90):
        for tag in ("TAG1", "TAG2", "TAG3"):
            for attempt in range(1, 6):
                rows.append([order, angle, tag, attempt, 1.5, 150, 150, "Papelao", 3, None, None, None, None])
                order += 1
    create_repeated_test_sheet(
        workbook, "Orientacao", headers, rows, [10, 11, 12, 13],
        [9, 18, 11, 10, 15, 18, 16, 21, 18, 16, 38, 24, 55], 10,
    )


def create_material(workbook: Workbook) -> None:
    headers = [
        "ordem", "material_suporte", "espacador_mm", "tag_id", "tentativa",
        "distancia_m", "altura_antena_cm", "altura_tag_cm", "orientacao_graus",
        "duracao_janela_s", "detectou_0_1", "epc_observado",
        "tempo_primeira_leitura_s", "observacoes",
    ]
    rows = []
    order = 1
    for material, spacer in (
        ("Papelao", 0),
        ("Plástico", 0),
        ("Madeira", 0),
        ("Vidro", 0),
        ("Metal direto", 0),
        ("Metal com espacador", 10),
    ):
        for attempt in range(1, 6):
            rows.append([order, material, spacer, "TAG1", attempt, 1.5, 150, 150, 0, 3, None, None, None, None])
            order += 1
    create_repeated_test_sheet(
        workbook, "Material", headers, rows, [11, 12, 13, 14],
        [9, 25, 16, 11, 10, 15, 18, 16, 18, 18, 16, 38, 24, 55], 11,
    )


def create_inventory(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Inventario")
    headers = [
        "rodada", "duracao_janela_s", "qtd_tags_presentes", "epcs_detectados_separados_por_virgula",
        "qtd_tags_detectadas", "cobertura_%", "completa_0_1", "leituras_externas",
        "percurso", "observacoes",
    ]
    sheet.append(headers)
    style_header(sheet)
    for round_number in range(1, 6):
        row = sheet.max_row + 1
        sheet.append([round_number, 25, 5, None, None, f'=IF(C{row}>0,E{row}/C{row}*100,"")', f'=IF(E{row}="","",--(E{row}=C{row}))', None, "Percurso curto padronizado", None])
        for column in (4, 5, 8, 9, 10):
            mark_input(sheet.cell(row, column))
        mark_formula(sheet.cell(row, 6))
        mark_formula(sheet.cell(row, 7))
        sheet.cell(row, 6).number_format = "0.0"
    add_binary_validation(sheet, "G2:G6")
    set_widths(sheet, [10, 20, 20, 58, 22, 16, 16, 18, 32, 55])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:J6"


def create_raw_log(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Log_Serial_Opcional")
    headers = ["timestamp", "ensaio", "condicao", "tentativa", "epc", "leitura_valida_0_1", "linha_serial_original", "observacoes"]
    sheet.append(headers)
    style_header(sheet)
    for row in range(2, 202):
        for column in range(1, 9):
            mark_input(sheet.cell(row, column))
    add_binary_validation(sheet, "F2:F201")
    set_widths(sheet, [24, 20, 28, 12, 42, 22, 80, 50])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:H201"


def create_summary(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Resumo", 2)
    style_title(sheet, "RESUMO AUTOMÁTICO — NÃO DIGITAR NAS CÉLULAS AZUIS", 7)
    headers = ["Ensaio", "Condição", "Planejado", "Preenchido", "Sucessos", "Taxa_%", "Situação"]
    sheet.append(headers)
    style_header(sheet, 2)
    row = 3
    for distance in (0.5, 1.0, 1.5, 2.0):
        sheet.cell(row, 1, "Alcance")
        sheet.cell(row, 2, distance)
        sheet.cell(row, 2).number_format = '0.0" m"'
        sheet.cell(row, 3, 15)
        sheet.cell(row, 4, f'=COUNTIFS(Alcance!$B:$B,B{row},Alcance!$J:$J,"<>")')
        sheet.cell(row, 5, f'=COUNTIFS(Alcance!$B:$B,B{row},Alcance!$J:$J,1)')
        sheet.cell(row, 6, f'=IF(D{row}>0,E{row}/D{row}*100,"")')
        sheet.cell(row, 7, f'=IF(D{row}=C{row},"COMPLETO","PENDENTE")')
        row += 1
    for angle in (0, 45, 90):
        sheet.cell(row, 1, "Orientacao")
        sheet.cell(row, 2, angle)
        sheet.cell(row, 2).number_format = '0"°"'
        sheet.cell(row, 3, 15)
        sheet.cell(row, 4, f'=COUNTIFS(Orientacao!$B:$B,B{row},Orientacao!$J:$J,"<>")')
        sheet.cell(row, 5, f'=COUNTIFS(Orientacao!$B:$B,B{row},Orientacao!$J:$J,1)')
        sheet.cell(row, 6, f'=IF(D{row}>0,E{row}/D{row}*100,"")')
        sheet.cell(row, 7, f'=IF(D{row}=C{row},"COMPLETO","PENDENTE")')
        row += 1
    for material in (
        "Papelao",
        "Plástico",
        "Madeira",
        "Vidro",
        "Metal direto",
        "Metal com espacador",
    ):
        sheet.cell(row, 1, "Material")
        sheet.cell(row, 2, material)
        sheet.cell(row, 3, 5)
        sheet.cell(row, 4, f'=COUNTIFS(Material!$B:$B,B{row},Material!$K:$K,"<>")')
        sheet.cell(row, 5, f'=COUNTIFS(Material!$B:$B,B{row},Material!$K:$K,1)')
        sheet.cell(row, 6, f'=IF(D{row}>0,E{row}/D{row}*100,"")')
        sheet.cell(row, 7, f'=IF(D{row}=C{row},"COMPLETO","PENDENTE")')
        row += 1
    sheet.cell(row, 1, "Inventario")
    sheet.cell(row, 2, "5 janelas")
    sheet.cell(row, 3, 5)
    sheet.cell(row, 4, '=COUNT(Inventario!$E$2:$E$6)')
    sheet.cell(row, 5, '=SUM(Inventario!$G$2:$G$6)')
    sheet.cell(row, 6, '=IFERROR(AVERAGE(Inventario!$F$2:$F$6),"")')
    sheet.cell(row, 7, f'=IF(D{row}=C{row},"COMPLETO","PENDENTE")')

    for data_row in range(3, row + 1):
        for column in range(4, 8):
            mark_formula(sheet.cell(data_row, column))
        sheet.cell(data_row, 6).number_format = "0.0"
    set_widths(sheet, [18, 28, 14, 14, 14, 14, 18])
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:G{row}"


def wilson_formula(success_cell: str, count_cell: str, lower: bool) -> str:
    sign = "-" if lower else "+"
    return (
        f'=IF({count_cell}=0,"",100*(({success_cell}/{count_cell}+1.96^2/(2*{count_cell})'
        f"{sign}1.96*SQRT(({success_cell}/{count_cell})*(1-{success_cell}/{count_cell})/"
        f"{count_cell}+1.96^2/(4*{count_cell}^2)))/(1+1.96^2/{count_cell})))"
    )


def create_factorial_tag_support(workbook: Workbook) -> None:
    """Cria o fatorial completo, futuro, com resultados manuais vazios."""
    sheet = workbook.create_sheet("Fatorial_Tag_Suporte")
    headers = [
        "ordem_execucao", "bloco", "montagem_id", "condicao_id", "comando_serial",
        "tag_id", "epc_esperado", "suporte_id", "material_suporte", "distancia_m",
        "orientacao_graus", "tentativa_celula", "altura_antena_cm", "altura_tag_cm",
        "duracao_janela_s", "detectou_0_1", "epc_observado",
        "tempo_primeira_leitura_s", "rssi_raw", "leituras_validas", "leituras_externas",
        "timeouts", "quadros_invalidos", "foto_montagem", "observacoes", "consistencia",
    ]
    sheet.append(headers)
    style_header(sheet)

    supports = [
        ("PAP1", "Papelao"),
        ("MAD1", "Madeira"),
        ("PLA1", "Plastico"),
    ]
    mountings = [
        (tag, support_id, material)
        for tag in ("TAG1", "TAG2", "TAG3")
        for support_id, material in supports
    ]
    seeded = random.Random(20260715)
    seeded.shuffle(mountings)

    order = 1
    for block in range(1, 4):
        offset = 3 * (block - 1)
        mounting_order = mountings[offset:] + mountings[:offset]
        geometries = [
            (distance, angle)
            for distance in (1.0, 2.0)
            for angle in (0, 45, 90)
        ]
        random.Random(20260715 + block).shuffle(geometries)
        for tag, support_id, material in mounting_order:
            mounting_id = f"B{block:02d}-{tag}-{support_id}"
            tag_number = int(tag[-1])
            for distance, angle in geometries:
                condition_id = (
                    f"B{block:02d}_{support_id}_D{int(distance * 10):02d}_A{angle:03d}"
                )
                row = sheet.max_row + 1
                command = f"RUN,FATORIAL,{condition_id},{tag_number},1"
                sheet.append([
                    order,
                    block,
                    mounting_id,
                    condition_id,
                    command,
                    tag,
                    f'=IFERROR(INDEX(Tags!$B$2:$B$6,MATCH(F{row},Tags!$A$2:$A$6,0)),"")',
                    support_id,
                    material,
                    distance,
                    angle,
                    block,
                    150,
                    150,
                    3,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    (
                        f'=IF(P{row}="","PENDENTE",IF(OR('
                        f'AND(P{row}=1,OR(Q{row}="",AND(G{row}<>"",Q{row}<>G{row}))),'
                        f'AND(P{row}=0,OR(Q{row}<>"",R{row}<>"",S{row}<>"",T{row}>0))'
                        f'),"REVISAR","OK"))'
                    ),
                ])
                for column in range(16, 26):
                    mark_input(sheet.cell(row, column))
                mark_formula(sheet.cell(row, 7))
                mark_formula(sheet.cell(row, 26))
                order += 1

    if order != 163:
        raise RuntimeError("o fatorial deve conter exatamente 162 linhas")

    add_header_comments(sheet, {
        "B1": "Bloco independente. Remova e reinstale a tag no suporte antes de iniciar cada montagem do bloco.",
        "C1": "Identifica uma montagem física única de TAG × suporte dentro de um bloco.",
        "D1": "Código ASCII estável da condição, usado também no log serial.",
        "E1": "Copie e envie um comando por linha. O último campo é 1 para produzir um registro individual.",
        "G1": "Fórmula de consulta à aba Tags; não digite por cima.",
        "H1": "Use cupons de mesmas dimensões e a mesma fita/espessura de fixação.",
        "P1": "1 somente se o EPC esperado aparecer na janela de 3 s; caso contrário, 0.",
        "S1": "Valor bruto, sem converter para dBm, salvo se a conversão estiver documentada.",
        "X1": "Nome da fotografia da montagem, se houver.",
        "Z1": "Verificação automática básica; REVISAR não substitui conferência do log.",
    })

    add_whole_validation(sheet, "B2:B163", 1, 3)
    add_list_validation(sheet, "F2:F163", ["TAG1", "TAG2", "TAG3"])
    add_list_validation(sheet, "H2:H163", ["PAP1", "MAD1", "PLA1"])
    add_list_validation(sheet, "I2:I163", ["Papelao", "Madeira", "Plastico"])
    add_list_validation(sheet, "J2:J163", [1, 2])
    add_list_validation(sheet, "K2:K163", [0, 45, 90])
    add_whole_validation(sheet, "L2:L163", 1, 3)
    add_binary_validation(sheet, "P2:P163")
    add_decimal_validation(sheet, "R2:R163", 0, 3)
    add_whole_validation(sheet, "S2:S163", 0, 255)
    for column in ("T", "U", "V", "W"):
        add_whole_validation(sheet, f"{column}2:{column}163", 0, 1_000_000)

    sheet.conditional_formatting.add(
        "P2:P163",
        CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor=RED)),
    )
    sheet.conditional_formatting.add(
        "P2:P163",
        CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor=GREEN)),
    )
    sheet.conditional_formatting.add(
        "Z2:Z163",
        CellIsRule(operator="equal", formula=['"REVISAR"'], fill=PatternFill("solid", fgColor=RED)),
    )
    sheet.conditional_formatting.add(
        "Z2:Z163",
        CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=GREEN)),
    )

    style_side_title(
        sheet,
        "AB1:AL1",
        "FATORIAL FUTURO / NÃO REALIZADO — METADADOS, INSTRUÇÕES E RESUMO",
    )
    metadata = [
        ("Status", "FUTURA / NÃO REALIZADA", "Não há resultados experimentais nesta aba."),
        ("Semente", 20260715, "Embaralhamento reproduzível; rotações contrabalançam os blocos."),
        ("Desenho", "3 × 3 × 2 × 3 × 3", "TAG × suporte × distância × ângulo × bloco = 162 janelas."),
        ("Remontagem", "Obrigatória por bloco", "Refaça a fixação de cada TAG × suporte em B1, B2 e B3."),
        ("Cupons", "Mesmas dimensões", "Centralize a tag e use a mesma fita e espessura."),
        ("Interpretação", "Amostra de cupons", "Um cupom por material não generaliza a toda a população do material."),
        ("Tempo estimado", "35–45 min", "Inclui remontagem, alinhamento e conferência."),
        ("Data da coleta", None, "Preencher somente quando a campanha for executada."),
        ("Ambiente/sala", None, "Registrar nome, área e alterações do ambiente."),
        ("Operador", None, "Pessoa que executou a campanha."),
        ("Firmware/commit", None, "Versão efetivamente carregada."),
        (
            "Antena/setpoint",
            None,
            "Modelo, ganho, polarização e potência configurada lida de volta; isso não mede potência conduzida.",
        ),
    ]
    for row, (label, value, instruction) in enumerate(metadata, start=2):
        sheet.cell(row, 28, label).font = Font(bold=True)
        sheet.cell(row, 29, value)
        sheet.merge_cells(start_row=row, start_column=30, end_row=row, end_column=38)
        sheet.cell(row, 30, instruction).alignment = Alignment(wrap_text=True, vertical="top")
        if value is None:
            mark_input(sheet.cell(row, 29))

    summary_row = 16
    summary_headers = [
        "tag_id", "material", "distancia_m", "angulo", "planejado",
        "preenchido", "deteccoes", "taxa_%", "ic95_inf_%", "ic95_sup_%", "situacao",
    ]
    for index, value in enumerate(summary_headers, start=28):
        sheet.cell(summary_row, index, value)
    style_header_range(sheet, summary_row, 28, 38)

    current = summary_row + 1
    for tag in ("TAG1", "TAG2", "TAG3"):
        for material in ("Papelao", "Madeira", "Plastico"):
            for distance in (1.0, 2.0):
                for angle in (0, 45, 90):
                    sheet.cell(current, 28, tag)
                    sheet.cell(current, 29, material)
                    sheet.cell(current, 30, distance)
                    sheet.cell(current, 31, angle)
                    sheet.cell(current, 32, 3)
                    sheet.cell(
                        current,
                        33,
                        f'=COUNTIFS($F$2:$F$163,AB{current},$I$2:$I$163,AC{current},'
                        f'$J$2:$J$163,AD{current},$K$2:$K$163,AE{current},$P$2:$P$163,"<>")',
                    )
                    sheet.cell(
                        current,
                        34,
                        f'=COUNTIFS($F$2:$F$163,AB{current},$I$2:$I$163,AC{current},'
                        f'$J$2:$J$163,AD{current},$K$2:$K$163,AE{current},$P$2:$P$163,1)',
                    )
                    sheet.cell(current, 35, f'=IF(AG{current}=0,"",AH{current}/AG{current}*100)')
                    sheet.cell(current, 36, wilson_formula(f"AH{current}", f"AG{current}", True))
                    sheet.cell(current, 37, wilson_formula(f"AH{current}", f"AG{current}", False))
                    sheet.cell(current, 38, f'=IF(AG{current}=AF{current},"COMPLETO","PENDENTE")')
                    for column in range(33, 39):
                        mark_formula(sheet.cell(current, column))
                    for column in (35, 36, 37):
                        sheet.cell(current, column).number_format = "0.0"
                    current += 1

    set_widths(sheet, [
        16, 9, 22, 30, 48, 10, 34, 12, 20, 13, 18, 17, 18,
        16, 18, 16, 36, 24, 13, 17, 18, 12, 18, 24, 52, 16,
    ])
    sheet.column_dimensions["AA"].width = 3
    for column, width in {
        "AB": 18, "AC": 20, "AD": 14, "AE": 12, "AF": 12, "AG": 14,
        "AH": 13, "AI": 13, "AJ": 13, "AK": 13, "AL": 15,
    }.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:Z163"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "70AD47"


def create_repetition_session2(workbook: Workbook) -> None:
    """Cria uma única sessão futura com os 140 registros individuais do protocolo."""
    sheet = workbook.create_sheet("Repeticao_Sessao2")
    headers = [
        "ordem_execucao", "sessao_id", "origem_ref", "ensaio", "condicao_id",
        "comando_serial", "tag_id", "epc_esperado", "tentativa_rodada", "distancia_m",
        "orientacao_graus", "material_suporte", "espacador_mm", "altura_antena_cm",
        "altura_tag_cm", "duracao_janela_s", "qtd_tags_presentes", "detectou_0_1",
        "epc_observado", "tempo_primeira_leitura_s", "rssi_raw", "leituras_validas",
        "epcs_detectados", "qtd_tags_detectadas", "cobertura_calculada_%",
        "completa_calculada_0_1", "leituras_externas", "timeouts", "quadros_invalidos",
        "fonte_alimentacao", "observacoes", "consistencia",
    ]
    sheet.append(headers)
    style_header(sheet)

    tag_support = {
        "TAG1": "Papelao",
        "TAG2": "Madeira",
        "TAG3": "Plastico",
    }
    order = 1

    def append_row(
        source: str,
        test: str,
        condition: str,
        command: str,
        tag: str | None,
        attempt: int,
        distance: float | None,
        angle: int | None,
        material: str | None,
        spacer: int | None,
        duration: int,
        present: int | None = None,
    ) -> None:
        nonlocal order
        row = sheet.max_row + 1
        sheet.append([
            order,
            "S2",
            source,
            test,
            condition,
            command,
            tag,
            f'=IF(G{row}="","",IFERROR(INDEX(Tags!$B$2:$B$6,MATCH(G{row},Tags!$A$2:$A$6,0)),""))',
            attempt,
            distance,
            angle,
            material,
            spacer,
            150 if test != "INVENTARIO" else None,
            150 if test != "INVENTARIO" else None,
            duration,
            present,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            (
                f'=IF(D{row}="INVENTARIO",'
                f'IF(OR(Q{row}="",X{row}=""),"",X{row}/Q{row}*100),"")'
            ),
            (
                f'=IF(D{row}="INVENTARIO",'
                f'IF(X{row}="","",--(X{row}=Q{row})),"")'
            ),
            None,
            None,
            None,
            None,
            None,
            (
                f'=IF(D{row}="INVENTARIO",'
                f'IF(X{row}="","PENDENTE",IF(OR(X{row}>Q{row},'
                f'AND(X{row}>0,W{row}="")),"REVISAR","OK")),'
                f'IF(R{row}="","PENDENTE",IF(OR('
                f'AND(R{row}=1,OR(S{row}="",AND(H{row}<>"",S{row}<>H{row}))),'
                f'AND(R{row}=0,OR(S{row}<>"",T{row}<>"",U{row}<>"",V{row}>0))'
                f'),"REVISAR","OK")))'
            ),
        ])
        mark_formula(sheet.cell(row, 8))
        mark_formula(sheet.cell(row, 25))
        mark_formula(sheet.cell(row, 26))
        mark_formula(sheet.cell(row, 32))
        if test == "INVENTARIO":
            for column in (23, 24, 27, 28, 29, 30, 31):
                mark_input(sheet.cell(row, column))
            for column in range(18, 23):
                sheet.cell(row, column).fill = PatternFill("solid", fgColor=GRAY)
        else:
            for column in (18, 19, 20, 21, 22, 27, 28, 29, 30, 31):
                mark_input(sheet.cell(row, column))
            for column in (23, 24):
                sheet.cell(row, column).fill = PatternFill("solid", fgColor=GRAY)
        order += 1

    source_row = 2
    for distance in (0.5, 1.0, 1.5, 2.0):
        distance_label = f"{distance:.1f}m"
        for tag in ("TAG1", "TAG2", "TAG3"):
            tag_number = int(tag[-1])
            for attempt in range(1, 6):
                append_row(
                    f"Alcance!linha {source_row}",
                    "ALCANCE",
                    f"ALC_{distance_label}",
                    f"RUN,ALCANCE,{distance_label},{tag_number},1",
                    tag,
                    attempt,
                    distance,
                    0,
                    tag_support[tag],
                    0,
                    3,
                )
                source_row += 1

    source_row = 2
    for angle in (0, 45, 90):
        for tag in ("TAG1", "TAG2", "TAG3"):
            tag_number = int(tag[-1])
            for attempt in range(1, 6):
                append_row(
                    f"Orientacao!linha {source_row}",
                    "ORIENTACAO",
                    f"ORI_{angle:03d}",
                    f"RUN,ORIENTACAO,{angle}graus,{tag_number},1",
                    tag,
                    attempt,
                    1.5,
                    angle,
                    tag_support[tag],
                    0,
                    3,
                )
                source_row += 1

    material_conditions = [
        ("Papelao", 0, "PAPELAO"),
        ("Plástico", 0, "PLASTICO"),
        ("Madeira", 0, "MADEIRA"),
        ("Vidro", 0, "VIDRO"),
        ("Metal direto", 0, "METAL_DIRETO"),
        ("Metal com espacador", 10, "METAL_ESPACADOR_10mm"),
    ]
    source_row = 2
    for material, spacer, serial_condition in material_conditions:
        condition = f"MAT_{serial_condition}"
        for attempt in range(1, 6):
            append_row(
                f"Material!linha {source_row}",
                "MATERIAL",
                condition,
                f"RUN,MATERIAL,{serial_condition},1,1",
                "TAG1",
                attempt,
                1.5,
                0,
                material,
                spacer,
                3,
            )
            source_row += 1

    for round_number in range(1, 6):
        append_row(
            f"Inventario!linha {round_number + 1}",
            "INVENTARIO",
            "INV_5tags",
            f"INV,{round_number},25,5",
            None,
            round_number,
            None,
            None,
            None,
            None,
            25,
            5,
        )

    if order != 141:
        raise RuntimeError("a repetição da sessão 2 deve conter exatamente 140 linhas")

    add_header_comments(sheet, {
        "B1": "Sessão futura S2. Não indica que a coleta já ocorreu.",
        "C1": "Referência da linha do protocolo original; nenhum resultado foi copiado.",
        "F1": "Comando individual: repetições=1. Para inventário, a linha já usa INV.",
        "H1": "Consulta automática à aba Tags; vazio nas rodadas de inventário.",
        "R1": "Nos ensaios individuais: 1 detectou o EPC esperado; 0 não detectou.",
        "W1": "Somente inventário: lista de EPCs separados por vírgula.",
        "X1": "Somente inventário: quantidade de tags distintas detectadas.",
        "Y1": "Fórmula de cobertura do inventário; fica visualmente vazia antes da coleta.",
        "Z1": "Fórmula de rodada completa; fica visualmente vazia antes da coleta.",
        "AF1": "Verificação automática básica, sem substituir a conferência do log.",
    })

    add_list_validation(sheet, "D2:D141", ["ALCANCE", "ORIENTACAO", "MATERIAL", "INVENTARIO"])
    add_list_validation(sheet, "G2:G141", ["TAG1", "TAG2", "TAG3", "TAG4", "TAG5"])
    add_binary_validation(sheet, "R2:R141")
    add_decimal_validation(sheet, "T2:T141", 0, 25)
    add_whole_validation(sheet, "U2:U141", 0, 255)
    add_whole_validation(sheet, "V2:V141", 0, 1_000_000)
    add_whole_validation(sheet, "X2:X141", 0, 5)
    for column in ("AA", "AB", "AC"):
        add_whole_validation(sheet, f"{column}2:{column}141", 0, 1_000_000)
    add_list_validation(
        sheet,
        "AD2:AD141",
        ["Fonte externa 5 V", "USB Arduino", "Outra/documentada"],
    )

    sheet.conditional_formatting.add(
        "R2:R141",
        CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor=RED)),
    )
    sheet.conditional_formatting.add(
        "R2:R141",
        CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor=GREEN)),
    )
    sheet.conditional_formatting.add(
        "AF2:AF141",
        CellIsRule(operator="equal", formula=['"REVISAR"'], fill=PatternFill("solid", fgColor=RED)),
    )

    style_side_title(
        sheet,
        "AH1:AO1",
        "SESSÃO 2 FUTURA / NÃO REALIZADA — METADADOS E COMPARAÇÃO COM S1",
    )
    for row, warning in (
        (
            2,
            "PRIORIDADE: repetir em outro dia no mesmo laboratório para medir estabilidade temporal.",
        ),
        (
            3,
            "Segundo ambiente é opcional. Trocar dia e ambiente ao mesmo tempo confunde os efeitos; "
            "para separá-los, não altere ambos simultaneamente.",
        ),
        (
            4,
            "GET_POWER apenas lê o setpoint interno. Potência conduzida e EIRP efetiva exigem instrumentação RF.",
        ),
    ):
        sheet.merge_cells(start_row=row, start_column=34, end_row=row, end_column=41)
        cell = sheet.cell(row, 34, warning)
        cell.font = Font(bold=True, color="9C0006")
        cell.fill = PatternFill("solid", fgColor=YELLOW)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.row_dimensions[row].height = 34

    metadata = [
        ("Status", "FUTURA / NÃO REALIZADA", "Nenhum resultado foi copiado da primeira sessão."),
        ("Sessão", "S2", "Use outro identificador se esta aba for clonada para S3."),
        ("Tipo de execução", "Mesmo laboratório / outro dia", "Opção prioritária para estabilidade temporal."),
        ("Data da coleta", None, "Obrigatório antes da primeira janela."),
        ("Hora início", None, "Hora local."),
        ("Hora fim", None, "Hora local."),
        ("Ambiente/sala/área", None, "Mesmo laboratório primeiro; detalhe qualquer mudança."),
        ("Operador", None, "Pessoa responsável pela sessão."),
        ("Pessoas/movimentação", None, "Quantidade e circulação no ambiente."),
        ("Metais próximos", None, "Bancada, armário e distâncias aproximadas."),
        ("Firmware/commit", None, "Versão efetivamente carregada."),
        ("Antena/modelo", None, "Modelo usado na S1 se a finalidade for repetibilidade."),
        ("Ganho/polarização", None, "Valores nominais e configuração efetiva."),
        ("Potência comandada (centidBm)", None, "Setpoint enviado ao módulo."),
        (
            "Potência configurada lida (centidBm)",
            None,
            "GET_POWER confirma o setpoint interno; não mede potência no conector.",
        ),
        ("Região/canais lidos", None, "Registrar quadros brutos ou NÃO VERIFICADO."),
        ("Fonte", None, "Fonte, tensão e corrente nominais."),
        ("Fotos/logs/ocorrências", None, "Arquivos, planta e qualquer desvio."),
    ]
    for row, (label, value, instruction) in enumerate(metadata, start=5):
        sheet.cell(row, 34, label).font = Font(bold=True)
        sheet.cell(row, 35, value)
        sheet.merge_cells(start_row=row, start_column=36, end_row=row, end_column=41)
        sheet.cell(row, 36, instruction).alignment = Alignment(wrap_text=True, vertical="top")
        if value is None:
            mark_input(sheet.cell(row, 35))

    summary_row = 25
    summary_headers = [
        "ensaio", "condicao", "planejado", "preenchido", "deteccoes/completas",
        "taxa/cobertura_%", "baseline_S1_%", "delta_pp",
    ]
    for index, value in enumerate(summary_headers, start=34):
        sheet.cell(summary_row, index, value)
    style_header_range(sheet, summary_row, 34, 41)

    strata: list[tuple[str, str, int, int]] = []
    strata.extend(
        ("ALCANCE", f"ALC_{distance:.1f}m", 15, baseline_row)
        for distance, baseline_row in zip((0.5, 1.0, 1.5, 2.0), range(3, 7))
    )
    strata.extend(
        ("ORIENTACAO", f"ORI_{angle:03d}", 15, baseline_row)
        for angle, baseline_row in zip((0, 45, 90), range(7, 10))
    )
    for (_, _, serial_condition), baseline_row in zip(material_conditions, range(10, 16)):
        strata.append(("MATERIAL", f"MAT_{serial_condition}", 5, baseline_row))
    strata.append(("INVENTARIO", "INV_5tags", 5, 16))

    for current, (test, condition, planned, baseline_row) in enumerate(
        strata,
        start=summary_row + 1,
    ):
        sheet.cell(current, 34, test)
        sheet.cell(current, 35, condition)
        sheet.cell(current, 36, planned)
        if test == "INVENTARIO":
            sheet.cell(
                current,
                37,
                '=COUNTIFS($D$2:$D$141,"INVENTARIO",$X$2:$X$141,"<>")',
            )
            sheet.cell(current, 38, '=SUMIFS($Z$2:$Z$141,$D$2:$D$141,"INVENTARIO")')
            sheet.cell(
                current,
                39,
                '=IFERROR(AVERAGEIFS($Y$2:$Y$141,$D$2:$D$141,"INVENTARIO"),"")',
            )
        else:
            sheet.cell(
                current,
                37,
                f'=COUNTIFS($E$2:$E$141,AI{current},$R$2:$R$141,"<>")',
            )
            sheet.cell(
                current,
                38,
                f'=COUNTIFS($E$2:$E$141,AI{current},$R$2:$R$141,1)',
            )
            sheet.cell(current, 39, f'=IF(AK{current}=0,"",AL{current}/AK{current}*100)')
        sheet.cell(current, 40, f"=Resumo!F{baseline_row}")
        sheet.cell(
            current,
            41,
            f'=IF(OR(AM{current}="",AN{current}=""),"",AM{current}-AN{current})',
        )
        for column in range(37, 42):
            mark_formula(sheet.cell(current, column))
        for column in (39, 40, 41):
            sheet.cell(current, column).number_format = "0.0"

    set_widths(sheet, [
        16, 11, 22, 15, 24, 46, 10, 34, 18, 13, 18, 22, 15, 18, 16, 18,
        20, 16, 38, 24, 13, 17, 54, 22, 22, 22, 18, 12, 18, 24, 54, 16,
    ])
    sheet.column_dimensions["AG"].width = 3
    for column, width in {
        "AH": 22, "AI": 25, "AJ": 14, "AK": 14,
        "AL": 20, "AM": 19, "AN": 17, "AO": 14,
    }.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:AF141"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "A5A5A5"


def create_antenna_comparison(workbook: Workbook) -> None:
    """Cria a comparação futura de duas antenas em 90 janelas individuais."""
    sheet = workbook.create_sheet("Comparacao_Antenas")
    headers = [
        "ordem_execucao", "bloco", "sequencia_antenas", "antena_id", "condicao_id",
        "comando_serial", "modelo", "polarizacao", "modo_operacao_porta",
        "ganho_nominal_dbi", "perda_cabo_db",
        "potencia_configurada_comandada_centidbm",
        "potencia_configurada_lida_centidbm", "eirp_nominal_calculada_dbm", "tag_id",
        "epc_esperado", "suporte_id", "material_suporte", "distancia_m",
        "orientacao_graus", "altura_antena_cm", "altura_tag_cm", "duracao_janela_s",
        "detectou_0_1", "epc_observado", "tempo_primeira_leitura_s", "rssi_raw",
        "leituras_validas", "leituras_externas", "timeouts", "quadros_invalidos",
        "foto_alinhamento", "observacoes", "consistencia",
    ]
    sheet.append(headers)
    style_header(sheet)

    antenna_a = ("ANTENA_A_LINEAR_APCA8090", "A")
    antenna_b = ("ANTENA_B_ALTERNATIVA", "B")
    block_sequences = [
        [antenna_a, antenna_b],
        [antenna_b, antenna_a],
        [antenna_a, antenna_b],
        [antenna_b, antenna_a],
    ]
    if random.Random(20260715).randrange(2) == 0:
        block_sequences.append([antenna_a, antenna_b])
    else:
        block_sequences.append([antenna_b, antenna_a])

    order = 1
    for block, sequence in enumerate(block_sequences, start=1):
        tag_angles = [
            (tag, angle)
            for tag in ("TAG1", "TAG2", "TAG3")
            for angle in (0, 45, 90)
        ]
        random.Random(20260715 + block).shuffle(tag_angles)
        sequence_label = f"{sequence[0][1]}→{sequence[1][1]}"
        for antenna_id, antenna_short in sequence:
            for tag, angle in tag_angles:
                row = sheet.max_row + 1
                tag_number = int(tag[-1])
                condition_id = f"B{block:02d}_ANT{antenna_short}_A{angle:03d}"
                command = f"RUN,ANTENA,{condition_id},{tag_number},1"
                sheet.append([
                    order,
                    block,
                    sequence_label,
                    antenna_id,
                    condition_id,
                    command,
                    f'=IFERROR(INDEX($AK$7:$AK$8,MATCH(D{row},$AJ$7:$AJ$8,0)),"")',
                    f'=IFERROR(INDEX($AL$7:$AL$8,MATCH(D{row},$AJ$7:$AJ$8,0)),"")',
                    f'=IFERROR(INDEX($AM$7:$AM$8,MATCH(D{row},$AJ$7:$AJ$8,0)),"")',
                    f'=IFERROR(INDEX($AN$7:$AN$8,MATCH(D{row},$AJ$7:$AJ$8,0)),"")',
                    f'=IFERROR(INDEX($AO$7:$AO$8,MATCH(D{row},$AJ$7:$AJ$8,0)),"")',
                    f'=IFERROR(INDEX($AP$7:$AP$8,MATCH(D{row},$AJ$7:$AJ$8,0)),"")',
                    f'=IFERROR(INDEX($AQ$7:$AQ$8,MATCH(D{row},$AJ$7:$AJ$8,0)),"")',
                    (
                        f'=IF(OR(J{row}="",K{row}="",M{row}=""),"",'
                        f'M{row}/100+J{row}-K{row})'
                    ),
                    tag,
                    f'=IFERROR(INDEX(Tags!$B$2:$B$6,MATCH(O{row},Tags!$A$2:$A$6,0)),"")',
                    "PADRAO_COMUM",
                    "Papelao",
                    1.5,
                    angle,
                    150,
                    150,
                    3,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    (
                        f'=IF(X{row}="","PENDENTE",IF(OR('
                        f'AND(X{row}=1,OR(Y{row}="",AND(P{row}<>"",Y{row}<>P{row}))),'
                        f'AND(X{row}=0,OR(Y{row}<>"",Z{row}<>"",AA{row}<>"",AB{row}>0))'
                        f'),"REVISAR","OK"))'
                    ),
                ])
                for column in range(24, 34):
                    mark_input(sheet.cell(row, column))
                for column in range(7, 15):
                    mark_formula(sheet.cell(row, column))
                mark_formula(sheet.cell(row, 16))
                mark_formula(sheet.cell(row, 34))
                order += 1

    if order != 91:
        raise RuntimeError("a comparação de antenas deve conter exatamente 90 linhas")

    add_header_comments(sheet, {
        "B1": "Cinco blocos; a ordem A/B alterna e o bloco 5 é definido pela semente registrada.",
        "C1": "Ordem de troca das antenas dentro do bloco.",
        "F1": "Comando individual. O firmware registra ANTENA como nome do ensaio e a condição ASCII.",
        "I1": "Para dual-port, informe Porta 1, Porta 2 ou o arranjo real. Uma porta isolada não é diversidade.",
        "J1": "Ganho nominal do fabricante, não medição realizada nesta campanha.",
        "K1": "Perda documentada do cabo/conectores usada somente no cálculo nominal.",
        "L1": "Setpoint enviado ao módulo em centésimos de dBm.",
        "M1": "GET_POWER lê de volta o setpoint interno; não mede potência no conector.",
        "N1": "Cálculo nominal: setpoint_lido/100 + ganho_nominal − perda_cabo.",
        "X1": "Resultado manual: 1 detectou o EPC esperado; 0 não detectou.",
        "AH1": "Verificação automática básica; confira também os logs e a ordem física.",
    })

    add_whole_validation(sheet, "B2:B91", 1, 5)
    add_list_validation(
        sheet,
        "D2:D91",
        ["ANTENA_A_LINEAR_APCA8090", "ANTENA_B_ALTERNATIVA"],
    )
    add_list_validation(sheet, "O2:O91", ["TAG1", "TAG2", "TAG3"])
    add_list_validation(sheet, "Q2:Q91", ["PADRAO_COMUM"])
    add_list_validation(sheet, "R2:R91", ["Papelao"])
    add_list_validation(sheet, "S2:S91", [1.5])
    add_list_validation(sheet, "T2:T91", [0, 45, 90])
    add_binary_validation(sheet, "X2:X91")
    add_decimal_validation(sheet, "Z2:Z91", 0, 3)
    add_whole_validation(sheet, "AA2:AA91", 0, 255)
    for column in ("AB", "AC", "AD", "AE"):
        add_whole_validation(sheet, f"{column}2:{column}91", 0, 1_000_000)

    sheet.conditional_formatting.add(
        "X2:X91",
        CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor=RED)),
    )
    sheet.conditional_formatting.add(
        "X2:X91",
        CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor=GREEN)),
    )
    sheet.conditional_formatting.add(
        "AH2:AH91",
        CellIsRule(operator="equal", formula=['"REVISAR"'], fill=PatternFill("solid", fgColor=RED)),
    )

    style_side_title(
        sheet,
        "AJ1:AR1",
        "COMPARAÇÃO FUTURA / NÃO REALIZADA — CONFIGURAÇÃO, AVISOS E RESUMO",
    )
    warnings = [
        (
            2,
            "DUAL-PORT: registre porta/modo. Uma porta isolada não constitui diversidade; cada porta "
            "deve ser condição separada, ou use chave/combinação RF documentada.",
        ),
        (
            3,
            "Compare preferencialmente com EIRP nominal equivalente, ajustando o setpoint para o ganho "
            "de cada antena. Com o mesmo setpoint, o ensaio compara sistemas completos, não só polarização.",
        ),
        (
            4,
            "GET_POWER confirma apenas o setpoint interno. Potência conduzida no conector e EIRP efetiva "
            "exigem instrumentação RF apropriada.",
        ),
    ]
    for row, warning in warnings:
        sheet.merge_cells(start_row=row, start_column=36, end_row=row, end_column=44)
        cell = sheet.cell(row, 36, warning)
        cell.font = Font(bold=True, color="9C0006")
        cell.fill = PatternFill("solid", fgColor=YELLOW)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.row_dimensions[row].height = 38

    config_headers = [
        "antena_id", "modelo", "polarizacao", "modo_porta", "ganho_dbi", "perda_cabo_db",
        "setpoint_comandado_centidbm", "setpoint_lido_centidbm", "eirp_nominal_dbm",
    ]
    for index, value in enumerate(config_headers, start=36):
        sheet.cell(6, index, value)
    style_header_range(sheet, 6, 36, 44)
    antenna_config = [
        [
            "ANTENA_A_LINEAR_APCA8090",
            "Airplux APCA8090",
            "Linear",
            "Porta unica",
            5,
            None,
            None,
            None,
        ],
        [
            "ANTENA_B_ALTERNATIVA",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ],
    ]
    for row, values in enumerate(antenna_config, start=7):
        for offset, value in enumerate(values, start=36):
            sheet.cell(row, offset, value)
            if offset > 36:
                mark_input(sheet.cell(row, offset))
        sheet.cell(
            row,
            44,
            f'=IF(OR(AN{row}="",AO{row}="",AQ{row}=""),"",AQ{row}/100+AN{row}-AO{row})',
        )
        mark_formula(sheet.cell(row, 44))
    add_list_validation(sheet, "AL7:AL8", ["Linear", "Circular", "Dual"])
    add_list_validation(
        sheet,
        "AM7:AM8",
        ["Porta unica", "Porta 1", "Porta 2", "Combinacao/diversidade"],
    )
    add_decimal_validation(sheet, "AN7:AN8", -20, 50)
    add_decimal_validation(sheet, "AO7:AO8", 0, 100)
    add_whole_validation(sheet, "AP7:AQ8", 500, 2600)

    summary_row = 10
    summary_headers = [
        "antena_id", "angulo", "planejado", "preenchido", "deteccoes",
        "taxa_%", "ic95_inf_%", "ic95_sup_%",
    ]
    for index, value in enumerate(summary_headers, start=36):
        sheet.cell(summary_row, index, value)
    style_header_range(sheet, summary_row, 36, 43)
    current = summary_row + 1
    for antenna_id, _ in (antenna_a, antenna_b):
        for angle in (0, 45, 90):
            sheet.cell(current, 36, antenna_id)
            sheet.cell(current, 37, angle)
            sheet.cell(current, 38, 15)
            sheet.cell(
                current,
                39,
                f'=COUNTIFS($D$2:$D$91,AJ{current},$T$2:$T$91,AK{current},$X$2:$X$91,"<>")',
            )
            sheet.cell(
                current,
                40,
                f'=COUNTIFS($D$2:$D$91,AJ{current},$T$2:$T$91,AK{current},$X$2:$X$91,1)',
            )
            sheet.cell(current, 41, f'=IF(AM{current}=0,"",AN{current}/AM{current}*100)')
            sheet.cell(current, 42, wilson_formula(f"AN{current}", f"AM{current}", True))
            sheet.cell(current, 43, wilson_formula(f"AN{current}", f"AM{current}", False))
            for column in range(39, 44):
                mark_formula(sheet.cell(current, column))
            for column in (41, 42, 43):
                sheet.cell(current, column).number_format = "0.0"
            current += 1

    global_row = 19
    global_headers = ["antena_id", "planejado", "preenchido", "deteccoes", "taxa_%", "situacao"]
    for index, value in enumerate(global_headers, start=36):
        sheet.cell(global_row, index, value)
    style_header_range(sheet, global_row, 36, 41)
    for current, (antenna_id, _) in enumerate((antenna_a, antenna_b), start=20):
        sheet.cell(current, 36, antenna_id)
        sheet.cell(current, 37, 45)
        sheet.cell(current, 38, f'=COUNTIFS($D$2:$D$91,AJ{current},$X$2:$X$91,"<>")')
        sheet.cell(current, 39, f'=COUNTIFS($D$2:$D$91,AJ{current},$X$2:$X$91,1)')
        sheet.cell(current, 40, f'=IF(AL{current}=0,"",AM{current}/AL{current}*100)')
        sheet.cell(current, 41, f'=IF(AL{current}=AK{current},"COMPLETO","PENDENTE")')
        for column in range(38, 42):
            mark_formula(sheet.cell(current, column))
        sheet.cell(current, 40).number_format = "0.0"
    sheet.cell(23, 36, "Delta B − A (pp)")
    sheet.cell(23, 37, '=IF(OR(AN20="",AN21=""),"",AN21-AN20)')
    mark_formula(sheet.cell(23, 37))
    sheet.cell(23, 37).number_format = "0.0"

    metadata = [
        ("Status", "FUTURA / NÃO REALIZADA", "Todos os campos de resultado estão vazios."),
        ("Semente", 20260715, "Ordem reproduzível; pares TAG × ângulo iguais para as duas antenas no bloco."),
        ("Distância", "1,5 m", "Centro de fase ao plano da tag."),
        ("Suporte comum", "Papelao padronizado", "Mesmas dimensões, fita e posição para as três tags."),
        ("Alinhamento", "Boresight marcado", "Reposicionar centro de fase e fotografar cada bloco."),
        ("Cabo", None, "Use o mesmo cabo e registre identificação/perda."),
        ("Data/ambiente", None, "Data, sala, pessoas e metais próximos."),
        ("Operador", None, "Pessoa responsável."),
        ("Firmware/commit", None, "Versão efetivamente carregada."),
        ("Instrumentação RF", None, "Equipamento/calibração se houver potência ou EIRP medida."),
    ]
    for row, (label, value, instruction) in enumerate(metadata, start=26):
        sheet.cell(row, 36, label).font = Font(bold=True)
        sheet.cell(row, 37, value)
        sheet.merge_cells(start_row=row, start_column=38, end_row=row, end_column=44)
        sheet.cell(row, 38, instruction).alignment = Alignment(wrap_text=True, vertical="top")
        if value is None:
            mark_input(sheet.cell(row, 37))

    set_widths(sheet, [
        16, 9, 18, 30, 27, 46, 24, 16, 23, 18, 16, 25, 23, 23, 10, 34, 18,
        20, 13, 18, 18, 16, 18, 16, 36, 24, 13, 17, 18, 12, 18, 24, 54, 16,
    ])
    sheet.column_dimensions["AI"].width = 3
    for column, width in {
        "AJ": 30, "AK": 22, "AL": 16, "AM": 18, "AN": 16,
        "AO": 16, "AP": 23, "AQ": 21, "AR": 20,
    }.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:AH91"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "ED7D31"


def main() -> int:
    args = parse_args()
    if OUTPUT.exists() and not args.overwrite:
        print(f"erro: {OUTPUT.name} já existe; use --overwrite para regenerar", file=sys.stderr)
        return 2

    workbook = Workbook()
    create_readme(workbook)
    create_day_plan(workbook)
    create_summary(workbook)
    create_configuration(workbook)
    create_tags(workbook)
    create_distance(workbook)
    create_orientation(workbook)
    create_material(workbook)
    create_inventory(workbook)
    create_raw_log(workbook)
    create_factorial_tag_support(workbook)
    create_repetition_session2(workbook)
    create_antenna_comparison(workbook)

    workbook.properties.title = "Coleta experimental mínima RFID UHF"
    workbook.properties.subject = "Planilha vazia para medições físicas do PG2"
    workbook.properties.description = (
        "Protocolo reduzido: alcance, orientação, material e inventário. "
        "Inclui modelos futuros de fatorial, repetição da sessão e comparação de antenas. "
        "Nenhum resultado experimental pré-preenchido."
    )
    workbook.properties.creator = "Projeto de Graduação 2"
    workbook.save(OUTPUT)

    # Confirma que todas as células de resultado permanecem vazias.
    check = load_workbook(OUTPUT, data_only=False)
    required_ranges = {
        "Alcance": [("J", 2, 61)],
        "Orientacao": [("J", 2, 46)],
        "Material": [("K", 2, 31)],
        "Inventario": [("E", 2, 6)],
        "Fatorial_Tag_Suporte": [
            (column, 2, 163) for column in ("P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y")
        ],
        "Repeticao_Sessao2": [
            (column, 2, 141)
            for column in ("R", "S", "T", "U", "V", "W", "X", "AA", "AB", "AC", "AD", "AE")
        ],
        "Comparacao_Antenas": [
            (column, 2, 91)
            for column in ("X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG")
        ],
    }
    for sheet_name, ranges in required_ranges.items():
        for column, first, last in ranges:
            if any(
                check[sheet_name][f"{column}{row}"].value is not None
                for row in range(first, last + 1)
            ):
                raise RuntimeError(
                    f"resultado pré-preenchido detectado em {sheet_name}!{column}{first}:{column}{last}"
                )

    expected_rows = {
        "Fatorial_Tag_Suporte": 163,
        "Repeticao_Sessao2": 141,
        "Comparacao_Antenas": 91,
    }
    expected_formulas = {
        "Fatorial_Tag_Suporte": 648,
        "Repeticao_Sessao2": 630,
        "Comparacao_Antenas": 941,
    }
    for sheet_name, expected_max_row in expected_rows.items():
        sheet = check[sheet_name]
        if sheet.max_row != expected_max_row:
            raise RuntimeError(
                f"contagem incorreta em {sheet_name}: {sheet.max_row - 1} linhas de dados"
            )
        formula_count = sum(
            cell.data_type == "f"
            for row in sheet.iter_rows()
            for cell in row
        )
        if formula_count != expected_formulas[sheet_name]:
            raise RuntimeError(
                f"fórmulas incorretas em {sheet_name}: {formula_count}"
            )

    print(f"Planilha criada: {OUTPUT}")
    print("Resultados pré-preenchidos: 0")
    print("Carga principal: 60 alcance + 45 orientação + 30 material + 5 inventários")
    print("Campanhas futuras: 162 fatorial + 140 sessão 2 + 90 comparação de antenas")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
