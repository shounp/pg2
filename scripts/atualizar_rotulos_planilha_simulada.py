#!/usr/bin/env python3
"""Atualiza somente os avisos textuais da planilha sintética usada no PG2.

O programa não modifica tentativas, detecções, RSSI, EPCs ou métricas. Ele torna
explícito que a base pode ser apresentada apenas como resultado de simulação.
"""

from __future__ import annotations

import argparse
from pathlib import Path
import sys

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "planilha_coleta_rfid_uhf_simulada.xlsx"


def find_row(sheet, label: str) -> int:
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value == label:
            return row
    raise ValueError(f"rótulo não encontrado em {sheet.title}: {label}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--write",
        action="store_true",
        help="grava os avisos revisados na planilha",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not WORKBOOK.is_file():
        print(f"erro: planilha não encontrada: {WORKBOOK}", file=sys.stderr)
        return 2

    workbook = load_workbook(WORKBOOK, data_only=False)
    readme = workbook["README"]
    configuration = workbook["Configuracao"]
    reference = workbook["Simulados_Referencia"]
    analysis = workbook["Analise_Erro"]

    readme.cell(find_row(readme, "Uso no TCC"), 2).value = (
        "Pode ser apresentada no TCC exclusivamente como resultado de simulação, "
        "com sua natureza sintética declarada no método, nas tabelas, figuras, "
        "resumo e conclusão. Não apresentar como medição física do protótipo."
    )
    configuration.cell(find_row(configuration, "Observações gerais"), 2).value = (
        "A base é usada na monografia como simulação exploratória. Uma coleta "
        "física futura será necessária para calibrar o modelo e validar o protótipo."
    )
    reference["B2"] = (
        "Não usar como resultado experimental. O uso acadêmico é permitido somente "
        "como saída sintética da simulação documentada."
    )
    analysis["A6"] = (
        "As faixas t de 95% descrevem apenas a dispersão interna entre configurações "
        "sintéticas; não são intervalos de confiança sobre o protótipo real."
    )
    analysis["L11"] = "Faixa t 95% inf. sucesso sintético (%)"
    analysis["M11"] = "Faixa t 95% sup. sucesso sintético (%)"
    analysis["N11"] = "Faixa t 95% inf. erro sintético (%)"
    analysis["O11"] = "Faixa t 95% sup. erro sintético (%)"

    workbook.properties.title = "Simulação de cenários RFID UHF — PG2"
    workbook.properties.subject = "Dados integralmente sintéticos"
    workbook.properties.description = (
        "Base RFID-UHF-SYN-20260712-v2. Não contém medições físicas do protótipo."
    )

    if not args.write:
        print("Rótulos verificados em memória; nenhum arquivo foi alterado.")
        return 0

    workbook.save(WORKBOOK)
    print(f"Avisos de simulação atualizados em {WORKBOOK.name}.")
    print("Nenhum valor quantitativo foi alterado.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
